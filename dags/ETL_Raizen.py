#!/usr/bin/env python
# coding: utf-8

##########################################################################################
#                                                                                        #
# etl pipeline to extract and structure the underlying data of two pivot tables:         #
#                                                                                        #
# -sales of oil derivative fuels by UF and product                                       #
# -sales of diesel by UF and type                                                        #
#                                                                                        #
# the totals of the extracted data must be equal to the totals of the pivot tables       #
#                                                                                        #
##########################################################################################

import os
import logging
import datetime
import pandas as pd
from urllib import request
from openpyxl import load_workbook
from airflow import DAG
from airflow.operators.python import PythonOperator


def creating_raw_data_dir() -> None:
    """ function which verify the existence of raw_data dir and create if it not exists """
    if os.path.isdir("./raw_data"):
        logging.info("Dir already exists")
        pass
    else:
        logging.info("Creating raw_data dir")
        os.system("mkdir ./raw_data")
        logging.info("Dir succesfully created")


def download_raw_data() -> None:
    """ function to extract the raw data and save into local directory """

    file_url = "https://github.com/raizen-analytics/data-engineering-test/raw/master/assets/vendas-combustiveis-m3.xls"
    file_xls = "./raw_data/vendas-combustiveis-m3.xls"
    logging.info("Downloading .xls raw data")
    request.urlretrieve(file_url , file_xls)
    logging.info("Succesfully downloaded .xls raw data")


def convert_xls_into_xlsx() -> None:
    """ functtion which convert .xls into .xlsx using libreoffice """
    logging.info("Converting file .xls into .xlsx using libreoffice")
    os.system("libreoffice --headless --invisible --convert-to xlsx ./raw_data/vendas-combustiveis-m3.xls --outdir ./raw_data")
    logging.info("File .xls succesfully converted into .xlsx")

def read_xlsx_file() -> object:
    """ method to read .xlsx file and return the object with workbooks """
    wb = load_workbook("./raw_data/vendas-combustiveis-m3.xlsx")
    #print(f"wb = {wb}")
    return wb

def get_specific_sheets(wb, list_all_sheets, specific_sheet) -> None:
    """ method which receive workbook object defined from .xlsx file and extract an specific sheet """
    for i in list_all_sheets:
        if i != specific_sheet:
            #print(wb[i])
            wb.remove(wb[i])
    
    wb.save(f"./{specific_sheet}.xlsx")

def extract_sheets() -> None:
    """ method which receive a list of sheets and extract them from workbook """
    list_specific_sheets = ["DPCache_m3", "DPCache_m3_2"]
    for i in list_specific_sheets:
        logging.info(f"Extracting sheet {i}")
        wb = read_xlsx_file()
        get_specific_sheets(wb, wb.sheetnames, i)
        logging.info (f"Sheet {i} succesfully extracted")

def data_transformation_and_load():
    """ function which apply the desired data transformation to extracted sheets """
    list_specific_sheets = ["DPCache_m3", "DPCache_m3_2"]

    for i in list_specific_sheets:
        logging.info(f"Processing data of sheet {i}")
        df = pd.read_excel(f"./{i}.xlsx")

        #taking only product
        df["product"] = df["COMBUSTÍVEL"].str.split(" \(").str[0]

        #taking only unit
        df["unit"] = df["COMBUSTÍVEL"].str.split(" \(").str[1].str.split("\)").str[0]

        #renaming "ESTADO" column
        df = df.rename(columns={"ESTADO":"uf"})

        #dropping obsolete columns
        df = df.drop(columns=["COMBUSTÍVEL", "REGIÃO", "TOTAL"])

        #creating column 'volume' with respective 'month'
        df = pd.melt(df, id_vars=["product", "uf", "ANO", "unit"], var_name=["month"], value_name="volume")

        months_names={"Jan":"01", "Fev":"02", "Mar":"03", "Abr":"04", "Mai":"05", "Jun":"06",
                      "Jul":"07", "Ago":"08", "Set":"09", "Out":"10", "Nov":"11", "Dez":"12"}

        df["month"] = df.month.replace(months_names)

        #creating 'year_month' column
        df["year_month"] = df["ANO"].astype(str) + "-" + df["month"].astype(str)

        df = df.drop(columns=["ANO", "month"])

        #creating 'created_at' column
        df["created_at"] = datetime.datetime.today().replace(microsecond=0)

        #completing nulls with '0'
        df = df.fillna(0)

        #defining data type of final dataframe
        df["product"] = df["product"].astype(str)
        df["uf"] = df["uf"].astype(str)
        df["unit"] = df["unit"].astype(str)
        df["volume"] = df["volume"].astype(float)
        df["year_month"] = df["year_month"].astype(str)
        df["created_at"] = pd.to_datetime(df["created_at"])

        print("\n")
        logging.info(f"Data of sheet {i} successfully processed")
        logging.info("Visual check: ")
        print(df.head())
        #print(df.count())

        logging.info("Saving processed data into parquet format")
        persist_path = f'{i}'
        df.to_parquet(persist_path, engine='fastparquet', partition_cols=['product', 'year_month'])
        logging.info("Data successfully saved")



with DAG('ETL_Raízen', start_date = datetime.datetime(2022,5,18),
           schedule_interval = '30 * * * *' , catchup = False) as dag:

    creating_raw_data_dir = PythonOperator(
        task_id = 'creating_raw_data_dir',
        python_callable = creating_raw_data_dir
    )

    download_raw_data = PythonOperator(
        task_id = 'download_raw_data',
        python_callable = download_raw_data
    )

    convert_xls_into_xlsx = PythonOperator(
        task_id = 'convert_xls_into_xlsx',
        python_callable = convert_xls_into_xlsx
    )

    extract_sheets = PythonOperator(
        task_id = 'extract_sheets',
        python_callable = extract_sheets
    )

    data_transformation_and_load = PythonOperator(
        task_id = 'data_transformation_and_load',
        python_callable = data_transformation_and_load
    )

    creating_raw_data_dir >> download_raw_data >> convert_xls_into_xlsx >> extract_sheets >> data_transformation_and_load